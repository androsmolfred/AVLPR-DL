import os
import sys
import logging
import threading
import time
import re
import uuid
from datetime import datetime

# ── Suppress Paddle / PaddleOCR logs ──────────────────────────────────────────
os.environ["GLOG_minloglevel"] = "3"
os.environ["FLAGS_logtostderr"] = "0"
os.environ["PADDLE_CPP_LOG_LEVEL"] = "3"
os.environ["PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK"] = "True"

from flask import Flask, jsonify, request, Response
from flask_cors import CORS
import cv2
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

from paddleocr import PaddleOCR
from ultralytics import YOLO

from statelga import (
    LGA_MAP,
    NIGERIA_STATES,
    get_state_from_lga_prefix,
    get_state_from_text
)

from live_feed import LiveFeed


# Suppress Python-level PaddleOCR logs
logging.getLogger("ppocr").setLevel(logging.ERROR)
logging.getLogger("paddle").setLevel(logging.ERROR)

# Add project root to Python path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

app = Flask(__name__)

CORS(app, resources={
    r"/api/*": {
        "origins": "*",
        "methods": ["GET", "POST", "OPTIONS"],
        "allow_headers": ["Content-Type"],
        "expose_headers": ["Content-Type"],
        "supports_credentials": False
    }
})

# ── Configuration ─────────────────────────────────────────────────────────────
DEVICE = "cpu"
LOG_FILE = "batch_log.xlsx"
OUTPUT_FOLDER = "test_results"
EVIDENCE_FOLDER = os.path.join(OUTPUT_FOLDER, "evidence_plates")

# Caps (prevents freeze)
YOLO_MAX_DIM = 1280
OCR_MAX_WIDTH = 640
OCR_MAX_HEIGHT = 320

# Evidence/log throttling (for live feed)
LIVE_LOG_COOLDOWN_SEC = 30  # seconds

# Limits to keep CPU steady
LIVE_FRAME_SKIP = 7        # run detection every Nth frame

# Excel/IO operations can slow things down, so only log occasionally in live feed.


# ── Initialize models ──────────────────────────────────────────────────────────
model_lock = threading.Lock()
ocr_lock = threading.Lock()

try:
    model = YOLO("best.pt").to(DEVICE)
    reader = PaddleOCR(use_textline_orientation=True, lang="en")
    print("Models loaded successfully")
except Exception as e:
    print(f"Warning: Could not load models: {e}")
    model = None
    reader = None


# ── Helpers ────────────────────────────────────────────────────────────────────
def setup_excel():
    if not os.path.exists(LOG_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Timestamp", "Image_Name", "Plate_Number", "State_of_Origin", "Country", "Confidence"])
        wb.save(LOG_FILE)


def resize_for_yolo(img):
    h, w = img.shape[:2]
    m = max(h, w)
    if m <= YOLO_MAX_DIM:
        return img
    scale = YOLO_MAX_DIM / m
    return cv2.resize(img, (int(w * scale), int(h * scale)), interpolation=cv2.INTER_AREA)


def resize_for_ocr(img):
    h, w = img.shape[:2]
    if w <= OCR_MAX_WIDTH and h <= OCR_MAX_HEIGHT:
        return img
    scale = min(OCR_MAX_WIDTH / w, OCR_MAX_HEIGHT / h)
    return cv2.resize(img, (max(1, int(w * scale)), max(1, int(h * scale))), interpolation=cv2.INTER_AREA)


def clean_and_format_plate(raw_text):
    clean = re.sub(r"[^A-Z0-9]", "", raw_text.upper())

    if len(clean) == 8:
        chars = list(clean)

        def fix_num(c):
            return {"O": "0", "I": "1", "S": "5", "B": "8", "Z": "2", "A": "4", "G": "6"}.get(c, c)

        def fix_let(c):
            return {"0": "O", "1": "I", "5": "S", "8": "B", "4": "A", "6": "G"}.get(c, c)

        # positions: AAA-000-XX
        chars[0] = fix_let(chars[0])
        chars[1] = fix_let(chars[1])
        chars[2] = fix_let(chars[2])

        chars[3] = fix_num(chars[3])
        chars[4] = fix_num(chars[4])
        chars[5] = fix_num(chars[5])

        chars[6] = fix_let(chars[6])
        chars[7] = fix_let(chars[7])

        return f"{''.join(chars[:3])}-{''.join(chars[3:6])}-{''.join(chars[6:])}"

    if len(clean) >= 5 and any(c.isdigit() for c in clean):
        return clean

    return None


def log_to_excel(timestamp, name, plate, state, confidence):
    try:
        wb = load_workbook(LOG_FILE)
        ws = wb.active
        ws.append([timestamp, name, plate, state, "Nigeria", round(float(confidence), 2)])

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        last_row = ws.max_row
        for col_idx in range(1, 7):
            ws.cell(row=last_row, column=col_idx).fill = green_fill

        wb.save(LOG_FILE)
    except PermissionError:
        # If Excel file is open somewhere, skip logging
        pass


def run_paddle_ocr(img_crop):
    """
    OCR ONCE per plate crop.
    Returns list of (text, score).
    """
    if reader is None:
        return []

    safe_img = resize_for_ocr(img_crop)

    with ocr_lock:
        result = reader.predict(safe_img)

    if not result:
        return []

    pairs = []
    for res in result:
        texts = res.get("rec_text", [])
        scores = res.get("rec_score", [])
        for t, s in zip(texts, scores):
            if t and str(t).strip():
                pairs.append((str(t).strip(), float(s)))

    return pairs


def detect_plate_from_crop(crop, label, do_logging=False):
    """
    Runs OCR ONCE over the crop.
    - picks best plate by prob
    - state by matching any OCR text to known states
    - fallback to LGA prefix based on first 3 letters of plate
    """
    # Slight upscale for readability but still capped by resize_for_ocr
    crop_up = cv2.resize(crop, None, fx=1.2, fy=1.2, interpolation=cv2.INTER_CUBIC)

    ocr_pairs = run_paddle_ocr(crop_up)

    detected_state = "Unknown"
    best_state_prob = -1.0

    best_plate = None
    best_plate_prob = -1.0

    for text, prob in ocr_pairs:
        # State detection from OCR text
        state_check = get_state_from_text(text)
        if state_check != "Unknown" and prob > best_state_prob:
            detected_state = state_check
            best_state_prob = prob

        # Plate detection from OCR text
        formatted = clean_and_format_plate(text)
        if formatted and prob > best_plate_prob:
            best_plate = formatted
            best_plate_prob = prob

    if detected_state == "Unknown" and best_plate:
        detected_state = get_state_from_lga_prefix(best_plate)

    if not best_plate:
        return None

    confidence_pct = round(best_plate_prob * 100, 2)

    if do_logging:
        # Save evidence + Excel log
        safe_name = re.sub(r"[^A-Z0-9]", "", best_plate)
        os.makedirs(EVIDENCE_FOLDER, exist_ok=True)
        evidence_path = os.path.join(EVIDENCE_FOLDER, f"EVIDENCE_{safe_name}_{label}.jpg")
        try:
            cv2.imwrite(evidence_path, crop)
        except Exception:
            pass

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_to_excel(timestamp, label, best_plate, detected_state, confidence_pct)

    return {
        "filename": label,
        "status": "processed",
        "plate_number": best_plate,
        "state_of_origin": detected_state,
        "confidence": confidence_pct,
        "message": "License plate detected successfully"
    }


def detect_plate_from_frame(frame, label, do_logging=False):
    """
    YOLO -> crop boxes -> OCR on crop -> choose best result.
    Returns (result_dict_or_none, evidence_crop_or_none)
    """
    if model is None or reader is None:
        return ({"error": "Models not loaded"}, None)

    img_yolo = resize_for_yolo(frame)

    with model_lock:
        yolo_results = model(img_yolo, conf=0.15, verbose=False, device=DEVICE)

    best = None
    best_prob = -1.0
    best_crop = None

    for result in yolo_results:
        boxes = result.boxes
        if boxes is None:
            continue

        for box in boxes:
            x1, y1, x2, y2 = map(int, box.xyxy[0])
            pad = 15
            y1_p = max(0, y1 - pad)
            y2_p = min(img_yolo.shape[0], y2 + pad)
            x1_p = max(0, x1 - pad)
            x2_p = min(img_yolo.shape[1], x2 + pad)

            crop = img_yolo[y1_p:y2_p, x1_p:x2_p]
            if crop.size <= 0:
                continue

            res = detect_plate_from_crop(crop, label, do_logging=False)
            if res and res.get("status") == "processed":
                if res["confidence"] > best_prob:
                    best_prob = res["confidence"]
                    best = res
                    best_crop = crop

    if best and do_logging:
        # If best is chosen, save evidence/log once
        _ = detect_plate_from_crop(best_crop, label, do_logging=True)

    return (best, best_crop)


# ── Live Feed callbacks ────────────────────────────────────────────────────────
def live_log_callback(result_dict, evidence_crop, label):
    """
    Called by LiveFeed when cooldown allows logging.
    """
    if result_dict is None or result_dict.get("status") != "processed":
        return
    if evidence_crop is None:
        return

    best_plate = result_dict.get("plate_number")
    detected_state = result_dict.get("state_of_origin")
    confidence = result_dict.get("confidence")

    safe_name = re.sub(r"[^A-Z0-9]", "", best_plate)
    os.makedirs(EVIDENCE_FOLDER, exist_ok=True)

    evidence_path = os.path.join(EVIDENCE_FOLDER, f"EVIDENCE_{safe_name}_{label}.jpg")
    try:
        cv2.imwrite(evidence_path, evidence_crop)
    except Exception:
        pass

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_to_excel(timestamp, label, best_plate, detected_state, confidence)


# Initialize live feed instance (after detect_fn exists)
def resize_frame_for_live(frame):
    # Keep frame small for YOLO in live mode
    return resize_for_yolo(frame)

# detect_fn signature required by LiveFeed:
# detect_fn(frame, label, do_logging=False) -> (result_dict, evidence_crop)
live_feed = LiveFeed(
    detect_fn=lambda frame, label, do_logging=False: detect_plate_from_frame(frame, label, do_logging=do_logging),
    resize_frame_fn=resize_frame_for_live,
    log_callback=live_log_callback,
    buffer_max=50,
    frame_skip=LIVE_FRAME_SKIP,
    detection_cooldown=0.0,
    log_cooldown=LIVE_LOG_COOLDOWN_SEC,
    stream_jpeg_quality=70,
)


# ── Routes ──────────────────────────────────────────────────────────────────────
@app.route("/")
def home():
    return jsonify({
        "message": "Welcome to AVLPRDL Backend API",
        "endpoints": {
            "/api/test": "Backend health check",
            "/api/process-image": "Upload images -> detect plates",
            "/api/process-video": "Upload video -> sample frames -> detect plates",
            "/api/dashboard": "Excel analytics",
            "/api/live/start": "Start webcam",
            "/api/live/stop": "Stop webcam",
            "/api/live/stream": "MJPEG stream",
            "/api/live/latest": "Latest detection",
            "/api/live-data": "Live detection buffer"
        }
    })


@app.route("/api/test")
def test():
    return jsonify({
        "status": "success",
        "message": "Backend is working correctly",
        "timestamp": datetime.now().isoformat(),
        "models_loaded": model is not None and reader is not None
    })


@app.route("/api/process-image", methods=["POST"])
def process_image():
    try:
        files = request.files.getlist("images")
        if not files:
            return jsonify({"error": "No files uploaded", "results": []}), 400

        valid_files = [f for f in files if f.filename and f.filename.strip() != ""]
        if not valid_files:
            return jsonify({"error": "No valid files selected", "results": []}), 400

        allowed_extensions = (".png", ".jpg", ".jpeg", ".gif", ".bmp")
        for file in valid_files:
            if not file.filename.lower().endswith(allowed_extensions):
                return jsonify({"error": f"Invalid file type for {file.filename}", "results": []}), 400

        os.makedirs("uploads", exist_ok=True)
        temp_files = []
        results = []

        try:
            for file in valid_files:
                unique_filename = f"{uuid.uuid4().hex}_{file.filename}"
                file_path = os.path.join("uploads", unique_filename)
                temp_files.append(file_path)
                file.save(file_path)

                # Detect with logging enabled for uploaded images
                img = cv2.imread(file_path)
                if img is None:
                    results.append({"filename": file.filename, "status": "error", "message": "Could not read image"})
                    continue

                label = file.filename
                res, _ = detect_plate_from_frame(img, label, do_logging=True)
                results.append(res if res else {
                    "filename": label,
                    "status": "no_plate_found",
                    "message": "No license plate detected"
                })

            return jsonify({
                "status": "success",
                "total_files": len(valid_files),
                "results": results
            }), 200

        finally:
            for p in temp_files:
                try:
                    if os.path.exists(p):
                        os.remove(p)
                except Exception:
                    pass

    except Exception as e:
        return jsonify({"error": str(e), "results": []}), 500


@app.route("/api/process-video", methods=["POST"])
def process_video():
    try:
        if "file" not in request.files:
            return jsonify({"error": "No file uploaded"}), 400
        file = request.files["file"]
        if file.filename == "":
            return jsonify({"error": "No file selected"}), 400
        if not file.filename.lower().endswith((".mp4", ".avi", ".mov", ".mkv")):
            return jsonify({"error": "Invalid file type"}), 400

        os.makedirs("uploads", exist_ok=True)
        file_path = os.path.join("uploads", file.filename)
        file.save(file_path)

        try:
            cap = cv2.VideoCapture(file_path)
            if not cap.isOpened():
                return jsonify({"error": "Could not open video file"}), 400

            results = []
            frame_count = 0
            processed_frames = 0
            frame_skip = 10
            max_processed = 100

            while cap.isOpened():
                ret, frame = cap.read()
                if not ret:
                    break

                frame_count += 1
                if frame_count % frame_skip != 0:
                    continue

                processed_frames += 1
                label = f"{file.filename}_frame_{frame_count}"

                res, _ = detect_plate_from_frame(frame, label, do_logging=True)
                if res:
                    results.append(res)

                if processed_frames >= max_processed:
                    break

            return jsonify({
                "filename": file.filename,
                "status": "processed",
                "total_frames": frame_count,
                "processed_frames": processed_frames,
                "results": results
            }), 200

        finally:
            try:
                os.remove(file_path)
            except Exception:
                pass

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/dashboard")
def dashboard():
    try:
        if not os.path.exists(LOG_FILE):
            return jsonify({"total": 0, "states": {}, "avg_confidence": 0, "recent": []})

        df = pd.read_excel(LOG_FILE)
        total = len(df)
        states = df["State_of_Origin"].value_counts().to_dict()
        avg_conf = round(df["Confidence"].astype(float).mean(), 2) if total > 0 else 0
        recent = df.tail(10).to_dict(orient="records")

        return jsonify({"total": total, "states": states, "avg_confidence": avg_conf, "recent": recent})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/live/start", methods=["POST"])
def live_start():
    data = request.get_json() or {}
    camera_index = int(data.get("camera_index", 0))

    r = live_feed.start(camera_index=camera_index, width=640, height=480, fps=15)
    return jsonify(r), (200 if r.get("status") == "success" else 400)


@app.route("/api/live/stop", methods=["POST"])
def live_stop():
    r = live_feed.stop()
    return jsonify(r)


@app.route("/api/live/stream")
def live_stream():
    return Response(
        live_feed.stream_generator(),
        mimetype="multipart/x-mixed-replace; boundary=frame"
    )


@app.route("/api/live/latest")
def live_latest():
    return jsonify(live_feed.get_latest())


@app.route("/api/live-data")
def live_data():
    # Compatible with old frontend concept: list of recent live detections
    return jsonify({
        "status": "success",
        "live_detections": live_feed.get_buffer(),
        "timestamp": datetime.now().isoformat()
    })


@app.route("/api/connection-test")
def connection_test():
    return jsonify({
        "status": "success",
        "message": "Backend connection successful",
        "server_time": datetime.now().isoformat(),
        "models_loaded": model is not None and reader is not None
    })


if __name__ == "__main__":
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    os.makedirs(EVIDENCE_FOLDER, exist_ok=True)
    setup_excel()

    print("Starting AVLPRDL Backend Server...")
    print("Models loaded successfully" if model and reader else "Warning: Models not loaded")
    print("Server running at http://localhost:5000")

    # threaded=True helps Flask handle requests while live worker runs
    app.run(debug=False, host="0.0.0.0", port=5000, threaded=True)